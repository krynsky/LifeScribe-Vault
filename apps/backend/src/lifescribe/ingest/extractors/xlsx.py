from __future__ import annotations

from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook

from lifescribe.ingest.extractors.base import ExtractionResult

_MAX_ROWS = 1000


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).replace("|", "\\|").replace("\n", " ")


class XlsxExtractor:
    mimes: ClassVar[tuple[str, ...]] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    NAME: ClassVar[str] = "xlsx"
    VERSION: ClassVar[str] = "0.1.0"

    def extract(self, path: Path) -> ExtractionResult:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        chunks: list[str] = []
        sheet_names: list[str] = []
        try:
            for ws in wb.worksheets:
                sheet_names.append(ws.title)
                chunks.append(f"## Sheet: {ws.title}\n")
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    chunks.append("(empty)\n")
                    continue
                header = [_cell_str(c) for c in rows[0]]
                lines = [
                    "| " + " | ".join(header) + " |",
                    "| " + " | ".join("---" for _ in header) + " |",
                ]
                data = rows[1:]
                truncated = len(data) > _MAX_ROWS
                if truncated:
                    data = data[:_MAX_ROWS]
                for row in data:
                    cells = [_cell_str(c) for c in row]
                    padded = cells + [""] * (len(header) - len(cells))
                    lines.append("| " + " | ".join(padded[: len(header)]) + " |")
                if truncated:
                    lines.append("")
                    lines.append(f"... truncated at {_MAX_ROWS} rows ({len(rows) - 1} total) ...")
                chunks.append("\n".join(lines) + "\n")
        finally:
            wb.close()

        return ExtractionResult(
            body_markdown="\n".join(chunks),
            extra_frontmatter={"sheet_names": sheet_names},
            extractor=f"{self.NAME}@{self.VERSION}",
            confidence=0.95,
        )
